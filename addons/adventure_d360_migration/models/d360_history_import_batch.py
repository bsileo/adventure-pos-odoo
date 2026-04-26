# -*- coding: utf-8 -*-

import base64
import io
import re
from collections import Counter, defaultdict
from datetime import datetime

from odoo import _, api, fields, models
from odoo.exceptions import UserError


def _xmlid_slug(value, max_len=80):
    text = (value or "").strip()
    slug = re.sub(r"[^a-zA-Z0-9_.-]+", "_", text)
    slug = re.sub(r"_+", "_", slug).strip("._") or "empty"
    return slug[:max_len]


def _digits_only(phone):
    return re.sub(r"\D+", "", phone or "")


def _normalize_email(email):
    return (email or "").strip().lower() or False


class AdventureD360HistoryImportBatch(models.Model):
    _name = "adventure.d360.history.import.batch"
    _description = "D360 historical transaction XLSX import batch"
    _order = "id desc"

    name = fields.Char(required=True)
    source_file = fields.Binary(string="Source XLSX", attachment=True)
    source_filename = fields.Char()
    source_checksum = fields.Char(string="SHA-256", readonly=True)
    dedupe_mode = fields.Selection(
        [
            ("keep_all", "Keep all rows (default)"),
            ("collapse_identical", "Collapse consecutive identical lines per invoice"),
        ],
        default="keep_all",
        required=True,
        help="With collapse mode, consecutive rows that are identical across key columns "
        "within the same invoice import as a single line (first row kept).",
    )
    state = fields.Selection(
        [
            ("draft", "Preview"),
            ("imported", "Imported"),
            ("cancelled", "Cancelled"),
        ],
        default="draft",
        required=True,
    )
    upload_notes = fields.Text()

    row_count = fields.Integer(string="Rows", readonly=True)
    invoice_count = fields.Integer(string="Invoices", readonly=True)
    customer_key_count = fields.Integer(
        string="Distinct D360 customers",
        readonly=True,
        help="Distinct Customer ID values in the file.",
    )
    product_match_preview = fields.Integer(
        string="Lines with product match",
        readonly=True,
    )
    product_unmatched_preview = fields.Integer(
        string="Lines without product match",
        readonly=True,
    )
    duplicate_candidate_count = fields.Integer(
        string="Duplicate-looking rows",
        readonly=True,
        help="Rows that exactly match another row on the same invoice (same fingerprint).",
    )
    negative_qty_row_count = fields.Integer(
        string="Return / negative qty rows",
        readonly=True,
    )
    skipped_duplicate_lines = fields.Integer(
        string="Skipped on import (dedupe)",
        readonly=True,
        help="Rows not imported because dedupe mode collapsed them.",
    )
    skipped_aggregate_rows = fields.Integer(
        string="Skipped (totals / non-lines)",
        readonly=True,
        help="Rows without Location and invoice number (e.g. report totals/footer), not imported.",
    )

    imported_at = fields.Datetime(readonly=True)
    order_ids = fields.One2many("adventure.history.order", "import_batch_id", string="Archive orders")

    def _require_openpyxl(self):
        try:
            import openpyxl  # noqa: F401
        except ImportError as exc:
            raise UserError(
                _(
                    "The Python package 'openpyxl' is required to read XLSX files. "
                    "Install it in the Odoo environment (see requirements.txt)."
                )
            ) from exc

    @api.model
    def _skip_non_transaction_row(self, entry):
        """Totals/footer rows: no Location or no invoice number (cannot archive-group)."""
        loc = str(entry.get("Location") or "").strip()
        inv = str(entry.get("Invoice number") or "").strip()
        return not (loc and inv)

    def _load_workbook_rows(self):
        """Return (rows, skipped_aggregate) where rows are dicts keyed by header; sheet order kept."""
        self.ensure_one()
        self._require_openpyxl()
        import openpyxl

        raw = base64.b64decode(self.source_file or b"")
        if not raw:
            raise UserError(_("The batch has no source file."))
        wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        try:
            ws = wb[wb.sheetnames[0]]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_row = next(rows_iter)
            except StopIteration:
                raise UserError(_("The spreadsheet is empty.")) from None
            headers = [(h or "").strip() if h is not None else "" for h in header_row]
            if not any(headers):
                raise UserError(_("The spreadsheet has no header row."))
            out = []
            skipped_aggregate = 0
            for sheet_row_index, row in enumerate(rows_iter, start=2):
                values = list(row)
                if all(v is None or str(v).strip() == "" for v in values):
                    continue
                entry = {"_sheet_row": sheet_row_index}
                for col_idx, key in enumerate(headers):
                    if not key:
                        continue
                    val = values[col_idx] if col_idx < len(values) else None
                    entry[key] = val
                if self._skip_non_transaction_row(entry):
                    skipped_aggregate += 1
                    continue
                out.append(entry)
            return out, skipped_aggregate
        finally:
            wb.close()

    @api.model
    def _row_fingerprint(self, row):
        """Stable tuple for duplicate detection within an invoice."""
        keys = (
            "Part number",
            "Barcode",
            "Description",
            "Serial Number",
            "Sold Qty",
            "Unit Price",
            "Ext. price",
            "Delivered Qty",
            "Returned Qty",
            "Cost",
            "Ext. cost",
            "Technician",
            "Instructor",
        )
        return tuple(str(row.get(k) if row.get(k) is not None else "").strip() for k in keys)

    def _validate_required_headers(self, sample_row_keys):
        required = (
            "Location",
            "Invoice number",
            "Customer ID",
            "Part number",
            "Barcode",
            "Description",
        )
        missing = [h for h in required if h not in sample_row_keys]
        if missing:
            raise UserError(
                _("Missing required columns: %s") % ", ".join(missing)
            )

    def _match_partner(self, row):
        Partner = self.env["res.partner"]
        cid = str(row.get("Customer ID") or "").strip()
        if cid:
            by_id = Partner.search([("d360_customer_id", "=", cid)], limit=2)
            if len(by_id) == 1:
                return by_id, "d360_id", ""
            if len(by_id) > 1:
                return Partner, "ambiguous", "Multiple partners share this D360 Customer ID."

        email = _normalize_email(row.get("Customer Email"))
        if email:
            found = Partner.search([("email", "=ilike", email)], limit=2)
            if len(found) == 1:
                return found, "email", ""
            if len(found) > 1:
                return Partner, "ambiguous", "Multiple partners share this email."

        phone_raw = str(row.get("Customer phone") or "").strip()
        digits = _digits_only(phone_raw)
        if len(digits) >= 10:
            tail = digits[-10:]
            found = Partner.search(
                ["|", ("phone", "ilike", tail), ("mobile", "ilike", tail)],
                limit=2,
            )
            if len(found) == 1:
                return found, "phone", ""
            if len(found) > 1:
                return Partner, "ambiguous", "Multiple partners share this phone pattern."

        first = str(row.get("First Name") or "").strip()
        last = str(row.get("Last Name") or "").strip()
        name_guess = " ".join(p for p in (first, last) if p).strip()
        zip_code = str(row.get("Zip") or "").strip()
        city = str(row.get("City") or "").strip()
        if name_guess and (zip_code or city):
            domain = [("name", "ilike", name_guess)]
            if zip_code:
                domain.append(("zip", "=", zip_code))
            else:
                domain.append(("city", "ilike", city))
            found = Partner.search(domain, limit=2)
            if len(found) == 1:
                return found, "name_address", ""
            if len(found) > 1:
                return Partner, "ambiguous", "Multiple partners matched name + address."

        return Partner, "unmatched", _("No confident partner match.")

    def _match_product(self, row):
        Product = self.env["product.product"]
        barcode = str(row.get("Barcode") or "").strip()
        if barcode:
            found = Product.search([("barcode", "=", barcode)])
            if len(found) == 1:
                return found, "barcode", ""
            if len(found) > 1:
                return Product, "ambiguous", "Multiple products share this barcode."

        part = str(row.get("Part number") or "").strip()
        description = str(row.get("Description") or "").strip()
        vendor = str(row.get("Vendor") or "").strip().lower()
        manufacturer = str(row.get("Manufacturer") or "").strip().lower()

        if part:
            found = Product.search([("default_code", "=", part)])
            if len(found) == 1:
                return found, "default_code", ""
            if len(found) > 1:
                narrowed = found
                if vendor:
                    narrowed = narrowed.filtered(
                        lambda p: vendor in (p.display_name or "").lower()
                        or vendor in (p.product_tmpl_id.name or "").lower()
                    )
                if manufacturer and len(narrowed) > 1:
                    narrowed = narrowed.filtered(
                        lambda p: manufacturer in (p.display_name or "").lower()
                        or manufacturer in (p.product_tmpl_id.name or "").lower()
                    )
                if description and len(narrowed) > 1:
                    snippet = description.lower()[:24]
                    narrowed = narrowed.filtered(
                        lambda p: snippet in (p.display_name or "").lower()
                        or snippet in (p.product_tmpl_id.name or "").lower()
                    )
                if len(narrowed) == 1:
                    return narrowed, "composite", "Resolved among multiple SKU matches."
                if len(narrowed) > 1:
                    return Product, "ambiguous", "Multiple SKU matches after composite hints."

        return Product, "unmatched", ""

    def _to_float(self, value):
        if value is None or value is False:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip().replace(",", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0

    def _to_bool(self, value):
        if isinstance(value, bool):
            return value
        text = str(value).strip().lower()
        return text in ("1", "true", "yes", "y", "t")

    def _to_datetime(self, value):
        if value is None or value is False or value == "":
            return False
        if isinstance(value, datetime):
            dt = value
            if dt.tzinfo:
                dt = dt.replace(tzinfo=None)
            return fields.Datetime.to_string(dt)
        text = str(value).strip()
        if not text:
            return False
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d"):
            try:
                dt = datetime.strptime(text[:26], fmt)
                return fields.Datetime.to_string(dt)
            except ValueError:
                continue
        return False

    def _compute_preview_stats(self, rows):
        if not rows:
            raise UserError(_("No data rows in the spreadsheet."))
        self._validate_required_headers(set(rows[0].keys()))

        invoices = set()
        customers = set()
        product_matched = 0
        product_unmatched = 0
        neg_qty = 0
        dup_candidates = 0

        by_invoice = defaultdict(list)
        for row in rows:
            loc = str(row.get("Location") or "").strip()
            inv = str(row.get("Invoice number") or "").strip()
            by_invoice[(loc, inv)].append(row)
            invoices.add((loc, inv))
            cid = str(row.get("Customer ID") or "").strip()
            if cid:
                customers.add(cid)

        for _inv_key, inv_rows in by_invoice.items():
            fps = [self._row_fingerprint(r) for r in inv_rows]
            for cnt in Counter(fps).values():
                if cnt > 1:
                    dup_candidates += cnt - 1

        for row in rows:
            _prod, status, _notes = self._match_product(row)
            if status in ("barcode", "default_code", "composite"):
                product_matched += 1
            else:
                product_unmatched += 1
            sold = self._to_float(row.get("Sold Qty"))
            ret = self._to_float(row.get("Returned Qty"))
            if sold < 0 or ret != 0:
                neg_qty += 1

        return {
            "row_count": len(rows),
            "invoice_count": len(invoices),
            "customer_key_count": len(customers),
            "product_match_preview": product_matched,
            "product_unmatched_preview": product_unmatched,
            "duplicate_candidate_count": dup_candidates,
            "negative_qty_row_count": neg_qty,
        }

    def action_analyze(self):
        for batch in self:
            rows, skipped_aggregate = batch._load_workbook_rows()
            stats = batch._compute_preview_stats(rows)
            batch.write(
                {**stats, "skipped_aggregate_rows": skipped_aggregate, "state": "draft"}
            )
        return True

    def action_import_archive(self):
        Order = self.env["adventure.history.order"]
        Line = self.env["adventure.history.order.line"]
        for batch in self:
            if batch.state == "cancelled":
                raise UserError(_("Cancelled batches cannot be imported."))
            rows, skipped_aggregate = batch._load_workbook_rows()
            batch._compute_preview_stats(rows)  # validate headers again

            by_invoice = defaultdict(list)
            for row in rows:
                loc = str(row.get("Location") or "").strip()
                inv = str(row.get("Invoice number") or "").strip()
                by_invoice[(loc, inv)].append(row)

            skipped = 0
            for (location, invoice_number), inv_rows in sorted(
                by_invoice.items(), key=lambda x: (x[0][0], x[0][1])
            ):
                inv_rows.sort(key=lambda r: r.get("_sheet_row", 0))
                header_row = inv_rows[0]
                partner, cust_status, cust_notes = batch._match_partner(header_row)
                order_xmlid = "d360.history_order.%s.%s" % (
                    _xmlid_slug(location),
                    _xmlid_slug(invoice_number),
                )
                name = "%s / %s" % (invoice_number or _("(no invoice)"), location or _("(no location)"))
                inv_dt = batch._to_datetime(header_row.get("Date"))

                order_vals = {
                    "name": name[:256],
                    "d360_source_xmlid": order_xmlid[:512],
                    "source_system": "D360",
                    "location": location or False,
                    "invoice_type": str(header_row.get("Invoice type") or "").strip() or False,
                    "invoice_number": invoice_number or False,
                    "invoice_date": inv_dt or False,
                    "sales_person": str(header_row.get("Sales Person") or "").strip() or False,
                    "d360_customer_id": str(header_row.get("Customer ID") or "").strip() or False,
                    "partner_id": partner.id if len(partner) == 1 else False,
                    "customer_match_status": cust_status,
                    "customer_match_notes": cust_notes or False,
                    "snapshot_first_name": str(header_row.get("First Name") or "").strip() or False,
                    "snapshot_last_name": str(header_row.get("Last Name") or "").strip() or False,
                    "snapshot_email": str(header_row.get("Customer Email") or "").strip() or False,
                    "snapshot_phone": str(header_row.get("Customer phone") or "").strip() or False,
                    "snapshot_address_1": str(header_row.get("Address 1") or "").strip() or False,
                    "snapshot_address_2": str(header_row.get("Address 2") or "").strip() or False,
                    "snapshot_city": str(header_row.get("City") or "").strip() or False,
                    "snapshot_state": str(header_row.get("State") or "").strip() or False,
                    "snapshot_zip": str(header_row.get("Zip") or "").strip() or False,
                    "snapshot_country": str(header_row.get("Country") or "").strip() or False,
                    "snapshot_customer_type": str(header_row.get("Customer type") or "").strip() or False,
                    "import_batch_id": batch.id,
                }

                existing = Order.search([("d360_source_xmlid", "=", order_xmlid)], limit=1)
                if existing:
                    existing.line_ids.unlink()
                    existing.write(order_vals)
                    order = existing
                else:
                    order = Order.create(order_vals)

                seq = 0
                prev_fp = None
                fp_counts = defaultdict(int)
                for r in inv_rows:
                    fp_counts[batch._row_fingerprint(r)] += 1

                for row in inv_rows:
                    fp = batch._row_fingerprint(row)
                    is_dup_candidate = fp_counts[fp] > 1

                    if (
                        batch.dedupe_mode == "collapse_identical"
                        and prev_fp is not None
                        and fp == prev_fp
                    ):
                        skipped += 1
                        continue
                    prev_fp = fp
                    seq += 1
                    line_xmlid = "d360.history_order_line.%s.%s.%s" % (
                        _xmlid_slug(location),
                        _xmlid_slug(invoice_number),
                        seq,
                    )
                    prod, prod_status, prod_notes = batch._match_product(row)
                    sold = batch._to_float(row.get("Sold Qty"))
                    ret = batch._to_float(row.get("Returned Qty"))

                    line_vals = {
                        "order_id": order.id,
                        "line_sequence": seq,
                        "d360_line_xmlid": line_xmlid[:512],
                        "part_number": str(row.get("Part number") or "").strip() or False,
                        "barcode": str(row.get("Barcode") or "").strip() or False,
                        "description": str(row.get("Description") or "").strip() or False,
                        "serial_number": str(row.get("Serial Number") or "").strip() or False,
                        "category": str(row.get("Category") or "").strip() or False,
                        "vendor": str(row.get("Vendor") or "").strip() or False,
                        "department": str(row.get("Department") or "").strip() or False,
                        "manufacturer": str(row.get("Manufacturer") or "").strip() or False,
                        "taxable": batch._to_bool(row.get("Taxable")),
                        "tax_collected_1": batch._to_float(row.get("Tax Collected1")),
                        "tax_collected_2": batch._to_float(row.get("Tax Collected2")),
                        "tax_collected_3": batch._to_float(row.get("Tax Collected3")),
                        "sold_qty": sold,
                        "unit_price": batch._to_float(row.get("Unit Price")),
                        "extended_price": batch._to_float(row.get("Ext. price")),
                        "delivered_qty": batch._to_float(row.get("Delivered Qty")),
                        "returned_qty": ret,
                        "date_delivered": batch._to_datetime(row.get("Date Delivered")),
                        "cost": batch._to_float(row.get("Cost")),
                        "extended_cost": batch._to_float(row.get("Ext. cost")),
                        "margin": batch._to_float(row.get("Margin")),
                        "technician": str(row.get("Technician") or "").strip() or False,
                        "instructor": str(row.get("Instructor") or "").strip() or False,
                        "primary_color": str(row.get("Primary Color") or "").strip() or False,
                        "secondary_color": str(row.get("Secondary Color") or "").strip() or False,
                        "accent_color": str(row.get("Accent Color") or "").strip() or False,
                        "size": str(row.get("Size") or "").strip() or False,
                        "product_id": prod.id if len(prod) == 1 else False,
                        "product_match_status": prod_status,
                        "product_match_notes": prod_notes or False,
                        "is_duplicate_candidate": is_dup_candidate,
                        "is_negative_qty": bool(sold < 0 or ret != 0),
                    }
                    Line.create(line_vals)

            stats = batch._compute_preview_stats(rows)
            batch.write(
                {
                    **stats,
                    "state": "imported",
                    "imported_at": fields.Datetime.now(),
                    "skipped_duplicate_lines": skipped,
                    "skipped_aggregate_rows": skipped_aggregate,
                }
            )
        return True

    def action_cancel(self):
        self.write({"state": "cancelled"})
        return True
